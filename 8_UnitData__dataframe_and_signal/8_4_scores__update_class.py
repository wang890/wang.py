"""
通过学号关联查询更新班级信息
从学生基本信息文件中查找班级信息，更新到成绩单文件中
使用类似数据库关联查询的方法实现
"""

import pandas as pd
from pathlib import Path
import os


def update_class_by_student_id(
    target_file_path,
    target_sheet_name,
    target_id_col,
    target_class_col,
    source_file_path,
    source_sheet_name,
    source_id_col,
    source_class_col,
    backup=True,
    font_color=None
):
    """
    通过学号关联查询更新班级信息
    
    参数:
        target_file_path: 目标文件路径（需要更新的Excel文件）
        target_sheet_name: 目标文件中的sheet名称
        target_id_col: 目标文件中学号列的列名或列索引（从0开始）
        target_class_col: 目标文件中班级列的列名或列索引（从0开始）
        source_file_path: 源文件路径（包含学生基本信息的Excel文件）
        source_sheet_name: 源文件中的sheet名称
        source_id_col: 源文件中学号列的列名或列索引（从0开始）
        source_class_col: 源文件中班级列的列名或列索引（从0开始）
        backup: 是否在更新前创建备份文件（默认True）
        font_color: 更新班级信息时使用的字体颜色，可以是颜色名称（如'red'）或RGB值（如'FF0000'），默认None表示不设置颜色
    
    返回:
        dict: 包含更新统计信息的字典
    """
    print(f"\n[信息] 开始更新班级信息")
    print(f"  目标文件: {target_file_path}")
    print(f"  源文件: {source_file_path}")
    
    # 检查文件是否存在
    if not os.path.exists(target_file_path):
        print(f"[错误] 目标文件不存在: {target_file_path}")
        return None
    
    if not os.path.exists(source_file_path):
        print(f"[错误] 源文件不存在: {source_file_path}")
        return None
    
    try:
        # 尝试使用pandas读取，如果失败则使用openpyxl
        try:
            # 读取目标文件
            print(f"\n[信息] 正在读取目标文件...")
            target_df = pd.read_excel(target_file_path, sheet_name=target_sheet_name, header=0)
            print(f"  目标文件行数: {len(target_df)}")
            print(f"  目标文件列名: {list(target_df.columns)}")
        except (ImportError, Exception) as e:
            # 如果pandas读取失败，使用openpyxl直接读取
            print(f"\n[信息] pandas读取失败，使用openpyxl读取目标文件...")
            print(f"  错误信息: {e}")
            from openpyxl import load_workbook
            wb = load_workbook(target_file_path, data_only=True)
            if target_sheet_name not in wb.sheetnames:
                print(f"[错误] 目标文件中不存在sheet: {target_sheet_name}")
                return None
            ws = wb[target_sheet_name]
            
            # 读取表头和数据
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # 跳过完全空的行
                    data.append(row)
            
            target_df = pd.DataFrame(data, columns=headers)
            print(f"  目标文件行数: {len(target_df)}")
            print(f"  目标文件列名: {list(target_df.columns)}")
        
        # 处理列索引或列名
        if isinstance(target_id_col, int):
            target_id_col_name = target_df.columns[target_id_col]
        else:
            target_id_col_name = target_id_col
        
        if isinstance(target_class_col, int):
            target_class_col_name = target_df.columns[target_class_col]
        else:
            target_class_col_name = target_class_col
        
        print(f"  学号列: {target_id_col_name}")
        print(f"  班级列: {target_class_col_name}")
        
        try:
            # 读取源文件
            print(f"\n[信息] 正在读取源文件...")
            source_df = pd.read_excel(source_file_path, sheet_name=source_sheet_name, header=0)
            print(f"  源文件行数: {len(source_df)}")
            print(f"  源文件列名: {list(source_df.columns)}")
        except (ImportError, Exception) as e:
            # 如果pandas读取失败，使用openpyxl直接读取
            print(f"\n[信息] pandas读取失败，使用openpyxl读取源文件...")
            print(f"  错误信息: {e}")
            from openpyxl import load_workbook
            wb = load_workbook(source_file_path, data_only=True)
            if source_sheet_name not in wb.sheetnames:
                print(f"[错误] 源文件中不存在sheet: {source_sheet_name}")
                return None
            ws = wb[source_sheet_name]
            
            # 读取表头和数据
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # 跳过完全空的行
                    data.append(row)
            
            source_df = pd.DataFrame(data, columns=headers)
            print(f"  源文件行数: {len(source_df)}")
            print(f"  源文件列名: {list(source_df.columns)}")
        
        # 处理列索引或列名
        if isinstance(source_id_col, int):
            source_id_col_name = source_df.columns[source_id_col]
        else:
            source_id_col_name = source_id_col
        
        if isinstance(source_class_col, int):
            source_class_col_name = source_df.columns[source_class_col]
        else:
            source_class_col_name = source_class_col
        
        print(f"  学号列: {source_id_col_name}")
        print(f"  班级列: {source_class_col_name}")
        
        # 检查列是否存在
        if target_id_col_name not in target_df.columns:
            print(f"[错误] 目标文件中不存在列: {target_id_col_name}")
            return None
        
        if target_class_col_name not in target_df.columns:
            print(f"[错误] 目标文件中不存在列: {target_class_col_name}")
            return None
        
        if source_id_col_name not in source_df.columns:
            print(f"[错误] 源文件中不存在列: {source_id_col_name}")
            return None
        
        if source_class_col_name not in source_df.columns:
            print(f"[错误] 源文件中不存在列: {source_class_col_name}")
            return None
        
        # 统计需要更新的行数（班级为空的行）
        # 判断班级是否为空：None、NaN、空字符串、只包含空格的字符串
        def is_class_empty(value):
            if pd.isna(value):
                return True
            if isinstance(value, str):
                return value.strip() == ''
            return False
        
        # 找出需要更新的行
        empty_class_mask = target_df[target_class_col_name].apply(is_class_empty)
        rows_to_update = target_df[empty_class_mask].copy()
        print(f"\n[信息] 需要更新班级的行数: {len(rows_to_update)}")
        
        if len(rows_to_update) == 0:
            print("[信息] 没有需要更新的行，退出")
            return {
                'total_rows': len(target_df),
                'rows_to_update': 0,
                'rows_updated': 0,
                'rows_not_found': 0,
                'updated_students': []
            }
        
        # 显示需要更新的学号
        if len(rows_to_update) > 0:
            print(f"  需要更新的学号: {rows_to_update[target_id_col_name].tolist()[:10]}")
            if len(rows_to_update) > 10:
                print(f"  ... 还有 {len(rows_to_update) - 10} 个学号")
        
        # 准备源数据：创建学号到班级的映射字典（类似数据库的索引）
        # 去除源数据中的重复学号，如果有重复，保留第一个
        source_mapping = source_df[[source_id_col_name, source_class_col_name]].copy()
        source_mapping = source_mapping.drop_duplicates(subset=[source_id_col_name], keep='first')
        source_mapping = source_mapping[source_mapping[source_id_col_name].notna()]  # 去除学号为空的记录
        
        # 创建学号到班级的字典映射
        id_to_class = dict(zip(
            source_mapping[source_id_col_name],
            source_mapping[source_class_col_name]
        ))
        
        print(f"\n[信息] 源文件中有效学号数量: {len(id_to_class)}")
        
        # 使用类似数据库LEFT JOIN的方式更新数据
        # 方法：通过merge实现关联查询
        rows_to_update_with_class = rows_to_update.merge(
            source_mapping,
            left_on=target_id_col_name,
            right_on=source_id_col_name,
            how='left',
            suffixes=('', '_source')
        )
        
        # 统计更新结果
        updated_count = 0
        not_found_count = 0
        updated_students = []
        # 记录需要设置颜色的DataFrame索引（原始索引，用于后续匹配）
        rows_to_color_indices = set()  # 记录哪些DataFrame索引对应的行需要设置颜色
        
        # 更新目标DataFrame中的班级信息
        for idx, row in rows_to_update.iterrows():
            student_id = row[target_id_col_name]
            
            # 在源数据中查找对应的班级
            if student_id in id_to_class:
                new_class = id_to_class[student_id]
                # 更新目标DataFrame
                target_df.at[idx, target_class_col_name] = new_class
                updated_count += 1
                # 记录这个DataFrame索引，后续写入Excel时需要设置颜色
                rows_to_color_indices.add(idx)
                updated_students.append({
                    '学号': student_id,
                    '原班级': row[target_class_col_name],
                    '新班级': new_class
                })
            else:
                not_found_count += 1
                print(f"  [警告] 学号 {student_id} 在源文件中未找到")
        
        print(f"\n[信息] 更新统计:")
        print(f"  成功更新: {updated_count} 行")
        print(f"  未找到: {not_found_count} 行")
        
        # 验证更新结果
        print(f"\n[信息] 验证更新结果...")
        still_empty = target_df[target_df[target_class_col_name].apply(is_class_empty)]
        print(f"  更新后仍为空的行数: {len(still_empty)}")
        
        if len(still_empty) > 0:
            print(f"  仍为空的学号: {still_empty[target_id_col_name].tolist()[:10]}")
        
        # 创建备份文件
        if backup:
            backup_path = str(target_file_path).replace('.xlsx', '_backup.xlsx')
            print(f"\n[信息] 创建备份文件: {backup_path}")
            try:
                # 使用文件直接复制，完全保留原文件的所有内容（包括格式、VBA等）
                from shutil import copyfile
                
                # 直接复制文件，确保完全一致
                copyfile(target_file_path, backup_path)
                print(f"  [成功] 备份文件已创建")
            except Exception as e:
                print(f"  [警告] 备份文件创建失败: {e}")
                import traceback
                traceback.print_exc()
        
        # 保存更新后的文件
        print(f"\n[信息] 正在保存更新后的文件...")
        # 使用openpyxl读取和写入，保留其他sheet
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border
        from copy import copy
        
        # 加载工作簿
        wb = load_workbook(target_file_path, data_only=False)  # data_only=False确保可以读取格式
        
        # 保存原有sheet的所有单元格信息（在删除前）
        original_ws = None
        original_cells = {}  # {(row, col): {'value': value, 'format': ...}}
        original_column_widths = {}  # {col_letter: width}
        original_row_heights = {}  # {row_num: height}
        if target_sheet_name in wb.sheetnames:
            original_ws = wb[target_sheet_name]
            # 复制所有单元格的值和格式
            print(f"  [信息] 正在复制原sheet的所有单元格格式...")
            for row in range(1, original_ws.max_row + 1):
                for col in range(1, original_ws.max_column + 1):
                    orig_cell = original_ws.cell(row=row, column=col)
                    cell_info = {
                        'value': orig_cell.value,
                        'number_format': orig_cell.number_format,
                        'alignment': copy(orig_cell.alignment) if orig_cell.alignment else None,
                        'font': copy(orig_cell.font) if orig_cell.font else None,
                        'fill': copy(orig_cell.fill) if orig_cell.fill else None,
                        'border': copy(orig_cell.border) if orig_cell.border else None,
                    }
                    original_cells[(row, col)] = cell_info
            
            # 保存列宽信息
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, original_ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in original_ws.column_dimensions:
                    col_dim = original_ws.column_dimensions[col_letter]
                    if col_dim.width:
                        original_column_widths[col_letter] = col_dim.width
            
            # 保存行高信息
            for row_idx in range(1, original_ws.max_row + 1):
                if row_idx in original_ws.row_dimensions:
                    row_dim = original_ws.row_dimensions[row_idx]
                    if row_dim.height:
                        original_row_heights[row_idx] = row_dim.height
            
            # 删除旧的sheet
            wb.remove(original_ws)
        
        # 创建新的sheet
        ws = wb.create_sheet(target_sheet_name)
        
        # 准备字体颜色（如果指定了）
        class_font = None
        if font_color:
            try:
                # 支持颜色名称（如'red'）或RGB值（如'FF0000'）
                if font_color.startswith('#'):
                    # 如果是以#开头的十六进制颜色，去掉#号
                    font_color = font_color[1:]
                elif font_color.lower() in ['red', 'blue', 'green', 'yellow', 'black', 'white']:
                    # 常见颜色名称映射
                    color_map = {
                        'red': 'FF0000',
                        'blue': '0000FF',
                        'green': '00FF00',
                        'yellow': 'FFFF00',
                        'black': '000000',
                        'white': 'FFFFFF'
                    }
                    font_color = color_map[font_color.lower()]
                # 创建字体对象
                class_font = Font(color=font_color)
                print(f"  [信息] 将使用字体颜色: {font_color}")
            except Exception as e:
                print(f"  [警告] 字体颜色设置失败: {e}，将不设置颜色")
        
        # 获取班级列和学号列在Excel中的列号
        headers = list(target_df.columns)
        class_col_num = None
        id_col_num = None
        for col_idx, header in enumerate(headers, start=1):
            if header == target_class_col_name:
                class_col_num = col_idx
            if header == target_id_col_name:
                id_col_num = col_idx
        
        # 先复制所有原有单元格的值和格式（完全保留原样）
        if original_cells:
            print(f"  [信息] 正在恢复原sheet的所有单元格...")
            for (row, col), cell_info in original_cells.items():
                cell = ws.cell(row=row, column=col)
                # 复制值（保持原始类型，特别是文本格式的数字）
                cell.value = cell_info['value']
                
                # 复制所有格式（完全保留原样）
                # number_format 总是存在，即使为空字符串也要设置
                cell.number_format = cell_info['number_format']
                if cell_info['alignment'] is not None:
                    cell.alignment = cell_info['alignment']
                if cell_info['font'] is not None:
                    cell.font = cell_info['font']
                if cell_info['fill'] is not None:
                    cell.fill = cell_info['fill']
                if cell_info['border'] is not None:
                    cell.border = cell_info['border']
        
        # 然后只更新需要更新的班级单元格
        print(f"  [信息] 正在更新班级信息...")
        for df_idx, df_original_idx in enumerate(target_df.index):
            if df_original_idx in rows_to_color_indices:
                excel_row_num = df_idx + 2  # Excel行号（+1是表头，+1是1-based）
                row_data = target_df.loc[df_original_idx]
                new_class_value = row_data[target_class_col_name]
                
                # 更新班级单元格
                cell = ws.cell(row=excel_row_num, column=class_col_num)
                cell.value = new_class_value
                
                # 保留原有格式，但更新字体颜色（如果需要）
                if (excel_row_num, class_col_num) in original_cells:
                    orig_cell_info = original_cells[(excel_row_num, class_col_num)]
                    # 保留原有格式（完全保留）
                    cell.number_format = orig_cell_info['number_format']
                    if orig_cell_info['alignment'] is not None:
                        cell.alignment = orig_cell_info['alignment']
                    if orig_cell_info['fill'] is not None:
                        cell.fill = orig_cell_info['fill']
                    if orig_cell_info['border'] is not None:
                        cell.border = orig_cell_info['border']
                    
                    # 更新字体：保留原有字体属性，只改变颜色
                    if class_font:
                        if orig_cell_info['font'] is not None:
                            new_font = copy(orig_cell_info['font'])
                            new_font.color = class_font.color
                            cell.font = new_font
                        else:
                            cell.font = class_font
                    elif orig_cell_info['font'] is not None:
                        cell.font = orig_cell_info['font']
                else:
                    # 如果没有原有格式，只设置字体颜色
                    if class_font:
                        cell.font = class_font
        
        # 恢复列宽
        if original_column_widths:
            print(f"  [信息] 正在恢复列宽...")
            for col_letter, width in original_column_widths.items():
                ws.column_dimensions[col_letter].width = width
        
        # 恢复行高
        if original_row_heights:
            print(f"  [信息] 正在恢复行高...")
            for row_idx, height in original_row_heights.items():
                ws.row_dimensions[row_idx].height = height
        
        # 保存文件
        wb.save(target_file_path)
        print(f"  [成功] 文件已保存")
        
        # 返回统计信息
        result = {
            'total_rows': len(target_df),
            'rows_to_update': len(rows_to_update),
            'rows_updated': updated_count,
            'rows_not_found': not_found_count,
            'rows_still_empty': len(still_empty),
            'updated_students': updated_students[:10]  # 只返回前10个示例
        }
        
        return result
        
    except Exception as e:
        print(f"[错误] 处理过程中出现异常: {e}")
        import traceback
        traceback.print_exc()
        return None


def verify_update_result(
    target_file_path,
    target_sheet_name,
    target_id_col,
    target_class_col,
    source_file_path,
    source_sheet_name,
    source_id_col,
    source_class_col
):
    """
    验证更新结果是否正确
    
    参数: 同update_class_by_student_id函数
    
    返回:
        dict: 验证结果
    """
    print(f"\n[信息] 开始验证更新结果...")
    
    try:
        # 尝试使用pandas读取，如果失败则使用openpyxl
        try:
            # 读取目标文件
            target_df = pd.read_excel(target_file_path, sheet_name=target_sheet_name, header=0)
        except (ImportError, Exception) as e:
            from openpyxl import load_workbook
            wb = load_workbook(target_file_path, data_only=True)
            if target_sheet_name not in wb.sheetnames:
                print(f"[错误] 目标文件中不存在sheet: {target_sheet_name}")
                return None
            ws = wb[target_sheet_name]
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            target_df = pd.DataFrame(data, columns=headers)
        
        try:
            # 读取源文件
            source_df = pd.read_excel(source_file_path, sheet_name=source_sheet_name, header=0)
        except (ImportError, Exception) as e:
            from openpyxl import load_workbook
            wb = load_workbook(source_file_path, data_only=True)
            if source_sheet_name not in wb.sheetnames:
                print(f"[错误] 源文件中不存在sheet: {source_sheet_name}")
                return None
            ws = wb[source_sheet_name]
            headers = [cell.value for cell in ws[1]]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            source_df = pd.DataFrame(data, columns=headers)
        
        # 处理列名
        if isinstance(target_id_col, int):
            target_id_col_name = target_df.columns[target_id_col]
        else:
            target_id_col_name = target_id_col
        
        if isinstance(target_class_col, int):
            target_class_col_name = target_df.columns[target_class_col]
        else:
            target_class_col_name = target_class_col
        
        if isinstance(source_id_col, int):
            source_id_col_name = source_df.columns[source_id_col]
        else:
            source_id_col_name = source_id_col
        
        if isinstance(source_class_col, int):
            source_class_col_name = source_df.columns[source_class_col]
        else:
            source_class_col_name = source_class_col
        
        # 创建源数据的学号到班级映射
        source_mapping = source_df[[source_id_col_name, source_class_col_name]].copy()
        source_mapping = source_mapping.drop_duplicates(subset=[source_id_col_name], keep='first')
        source_mapping = source_mapping[source_mapping[source_id_col_name].notna()]
        id_to_class = dict(zip(
            source_mapping[source_id_col_name],
            source_mapping[source_class_col_name]
        ))
        
        # 验证：检查目标文件中所有非空班级是否与源文件一致
        verification_results = {
            'total_rows': len(target_df),
            'matched_rows': 0,
            'mismatched_rows': 0,
            'empty_rows': 0,
            'not_found_rows': 0,
            'mismatches': []
        }
        
        def is_class_empty(value):
            if pd.isna(value):
                return True
            if isinstance(value, str):
                return value.strip() == ''
            return False
        
        for idx, row in target_df.iterrows():
            student_id = row[target_id_col_name]
            target_class = row[target_class_col_name]
            
            if is_class_empty(target_class):
                verification_results['empty_rows'] += 1
            elif student_id in id_to_class:
                source_class = id_to_class[student_id]
                # 比较班级（转换为字符串进行比较）
                if str(target_class).strip() == str(source_class).strip():
                    verification_results['matched_rows'] += 1
                else:
                    verification_results['mismatched_rows'] += 1
                    verification_results['mismatches'].append({
                        '学号': student_id,
                        '目标文件班级': target_class,
                        '源文件班级': source_class
                    })
            else:
                verification_results['not_found_rows'] += 1
        
        print(f"\n[验证结果]")
        print(f"  总行数: {verification_results['total_rows']}")
        print(f"  匹配的行数: {verification_results['matched_rows']}")
        print(f"  不匹配的行数: {verification_results['mismatched_rows']}")
        print(f"  仍为空的行数: {verification_results['empty_rows']}")
        print(f"  源文件中未找到的行数: {verification_results['not_found_rows']}")
        
        if verification_results['mismatches']:
            print(f"\n  不匹配的示例（前5个）:")
            for mismatch in verification_results['mismatches'][:5]:
                print(f"    学号 {mismatch['学号']}: 目标文件={mismatch['目标文件班级']}, 源文件={mismatch['源文件班级']}")
        
        return verification_results
        
    except Exception as e:
        print(f"[错误] 验证过程中出现异常: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == '__main__':
    # 设置路径
    base_dir = Path(__file__).parent
    
    # 目标文件：需要更新的成绩单文件
    target_file = base_dir / 'scores4_xlsx_items' / 'grd23yr25__张金平.xlsx'
    target_sheet = 'copy'
    target_id_col = 2  # 第2列（索引从0开始，所以是1）
    target_class_col = 3  # 第3列（索引从0开始，所以是2）
    
    # 源文件：学生基本信息文件
    source_file = base_dir / 'scores1_raw' / '2023级学生基本信息 (1).xlsx'
    # source_sheet = '2022级学生基本信息'
    source_sheet = '学生基本信息'
    source_id_col = 0  # 第1列（索引从0开始，所以是0）
    source_class_col = 8  # 第9列（索引从0开始，所以是8）
    
    # 执行更新
    print("=" * 60)
    print("开始更新班级信息")
    print("=" * 60)
    
    result = update_class_by_student_id(
        target_file_path=str(target_file),
        target_sheet_name=target_sheet,
        target_id_col=target_id_col,
        target_class_col=target_class_col,
        source_file_path=str(source_file),
        source_sheet_name=source_sheet,
        source_id_col=source_id_col,
        source_class_col=source_class_col,
        backup=True,
        font_color='blue'  # 设置更新的班级信息为红色
    )
    
    if result:
        print("\n" + "=" * 60)
        print("更新完成")
        print("=" * 60)
        print(f"\n更新统计:")
        print(f"  总行数: {result['total_rows']}")
        print(f"  需要更新的行数: {result['rows_to_update']}")
        print(f"  成功更新的行数: {result['rows_updated']}")
        print(f"  未找到的行数: {result['rows_not_found']}")
        print(f"  仍为空的行数: {result['rows_still_empty']}")
        
        # 执行验证
        print("\n" + "=" * 60)
        print("开始验证更新结果")
        print("=" * 60)
        
        verify_result = verify_update_result(
            target_file_path=str(target_file),
            target_sheet_name=target_sheet,
            target_id_col=target_id_col,
            target_class_col=target_class_col,
            source_file_path=str(source_file),
            source_sheet_name=source_sheet,
            source_id_col=source_id_col,
            source_class_col=source_class_col
        )
        
        if verify_result:
            print("\n验证完成")
    else:
        print("\n更新失败，请检查错误信息")

